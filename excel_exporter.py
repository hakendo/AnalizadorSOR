from __future__ import annotations
import os
from datetime import date
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

import history as hist_mod

# ── Palette ───────────────────────────────────────────────────────────────
_BLUE_HDR    = "1F4E79"
_BLUE_SUB    = "2E75B6"
_BLUE_LIGHT  = "BDD7EE"
_GRAY        = "F2F2F2"
_WHITE       = "FFFFFF"
_RED_LIGHT   = "FFD7D7"
_ORANGE_LIGHT = "FFE8CC"
_GREEN_DELTA = "E2EFDA"
_RED_DELTA   = "FCE4D6"

_DIR_LABELS = {'normal': 'Bidireccional', 'corta': 'Corta', 'larga': 'Larga'}


# ── Column registry ───────────────────────────────────────────────────────
# Each entry: key, label, width, num_format, source, field, thr_key, default_on
# source: 'event' | 'fiber' | 'fiber_first'  (fiber_first = show only on first event row)

ALL_DETAIL_COLS: list[dict] = [
    {'key': 'fibra_num',             'label': "Fibra N°",               'width': 9,
     'fmt': None,      'source': 'fiber_first', 'field': 'fibra_num',
     'thr_key': None,                    'default': True,  'required': True},
    {'key': 'direction',             'label': "Dirección",              'width': 13,
     'fmt': None,      'source': 'fiber_first', 'field': 'direction_label',
     'thr_key': None,                    'default': True,  'required': False},
    {'key': 'n_evento',              'label': "N° Evento",              'width': 9,
     'fmt': None,      'source': 'event',        'field': 'n_evento',
     'thr_key': None,                    'default': True,  'required': False},
    {'key': 'posicion_km',           'label': "Posición (km)",          'width': 13,
     'fmt': "0.0000",  'source': 'event',        'field': 'posicion_km',
     'thr_key': None,                    'default': True,  'required': True},
    {'key': 'longitud_intervalo_km', 'label': "Long. Intervalo (km)",   'width': 16,
     'fmt': "0.0000",  'source': 'event',        'field': 'longitud_intervalo_km',
     'thr_key': None,                    'default': True,  'required': False},
    {'key': 'perdida_intervalo_db',  'label': "Pérd. Intervalo (dB)",   'width': 16,
     'fmt': "0.0000",  'source': 'event',        'field': 'perdida_intervalo_db',
     'thr_key': 'perdida_intervalo_db',  'default': True,  'required': False},
    {'key': 'perdida_promedio_dbkm', 'label': "Pérd. Prom. (dB/km)",   'width': 16,
     'fmt': "0.0000",  'source': 'event',        'field': 'perdida_promedio_dbkm',
     'thr_key': 'perdida_promedio_dbkm', 'default': True,  'required': False},
    {'key': 'perdida_union_db',      'label': "Pérd. Unión (dB)",       'width': 14,
     'fmt': "0.0000",  'source': 'event',        'field': 'perdida_union_db',
     'thr_key': 'perdida_union_db',      'default': True,  'required': False},
    {'key': 'union_prom',            'label': "Pérd. Unión Prom. (dB)", 'width': 18,
     'fmt': "0.0000",  'source': 'fiber_first',  'field': 'perdida_union_promedio_db',
     'thr_key': None,                    'default': True,  'required': False},
    {'key': 'union_max',             'label': "Pérd. Unión Máx. (dB)",  'width': 18,
     'fmt': "0.0000",  'source': 'fiber_first',  'field': 'perdida_union_maxima_db',
     'thr_key': 'perdida_union_db',      'default': True,  'required': False},
]

ALL_SUMMARY_COLS: list[dict] = [
    {'key': 'fibra_num',               'label': "Fibra N°",               'width': 9,
     'fmt': None,       'field': 'fibra_num',                  'thr_key': None,
     'default': True,   'required': True},
    {'key': 'direction',               'label': "Dirección",              'width': 13,
     'fmt': None,       'field': 'direction_label',            'thr_key': None,
     'default': True,   'required': False},
    {'key': 'longitud_total_km',       'label': "Long. Total (km)",       'width': 14,
     'fmt': "0.0000",   'field': 'longitud_total_km',          'thr_key': None,
     'default': True,   'required': False},
    {'key': 'perdida_total_db',        'label': "Pérd. Total (dB)",       'width': 14,
     'fmt': "0.0000",   'field': 'perdida_total_db',           'thr_key': None,
     'default': True,   'required': False},
    {'key': 'perdida_promedio_dbkm',   'label': "Pérd. Prom. (dB/km)",   'width': 16,
     'fmt': "0.0000",   'field': 'perdida_promedio_dbkm',      'thr_key': 'perdida_promedio_dbkm',
     'default': True,   'required': False},
    {'key': 'perdida_union_promedio_db','label': "Pérd. Unión Prom. (dB)",'width': 18,
     'fmt': "0.0000",   'field': 'perdida_union_promedio_db',  'thr_key': None,
     'default': True,   'required': False},
    {'key': 'perdida_union_maxima_db', 'label': "Pérd. Unión Máx. (dB)", 'width': 18,
     'fmt': "0.0000",   'field': 'perdida_union_maxima_db',    'thr_key': 'perdida_union_db',
     'default': True,   'required': False},
    {'key': 'n_empalmes',              'label': "N° Empalmes",            'width': 12,
     'fmt': None,       'field': 'n_empalmes',                 'thr_key': None,
     'default': True,   'required': False},
    {'key': 'delta_union_max',         'label': "∆ Pérd. Unión Máx.",    'width': 16,
     'fmt': "+0.0000",  'field': 'delta_union_max',            'thr_key': None,
     'default': True,   'required': False},
    {'key': 'date',                    'label': "Fecha",                  'width': 12,
     'fmt': None,       'field': 'date',                       'thr_key': None,
     'default': True,   'required': False},
]


def default_column_config() -> dict:
    return {
        'detail':  [c['key'] for c in ALL_DETAIL_COLS  if c['default']],
        'summary': [c['key'] for c in ALL_SUMMARY_COLS if c['default']],
    }


def active_cols(registry: list[dict], enabled_keys: list[str]) -> list[dict]:
    key_set = set(enabled_keys)
    return [c for c in registry if c['key'] in key_set or c.get('required')]


# ── Style helpers ─────────────────────────────────────────────────────────

def _side() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row: int, col: int, value: Any = None,
          bold: bool = False, bg: str | None = None,
          align: str = "center", fmt: str | None = None,
          font_color: str = "000000") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, size=10, color=font_color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    c.border = _side()


def _thr_bg(value: float | None, threshold: float | None) -> str | None:
    if value is None or threshold is None:
        return None
    if value > threshold:
        return _RED_LIGHT
    if value > threshold * 0.8:
        return _ORANGE_LIGHT
    return None


# ── Value extractors ──────────────────────────────────────────────────────

def _detail_value(col: dict, fiber: dict, ev: dict, ei: int) -> Any:
    src = col['source']
    fld = col['field']
    if src == 'event':
        return ev.get(fld)
    if src == 'fiber_first':
        if ei > 0:
            return None
        if fld == 'direction_label':
            return _DIR_LABELS.get(fiber.get('direction', 'normal'), fiber.get('direction'))
        return fiber.get(fld)
    if src == 'fiber':
        if fld == 'direction_label':
            return _DIR_LABELS.get(fiber.get('direction', 'normal'), fiber.get('direction'))
        return fiber.get(fld)
    return None


def _detail_bg(col: dict, fiber: dict, ev: dict, ei: int,
               thr: dict, row_bg: str) -> str:
    key = col['key']
    thr_key = col.get('thr_key')
    is_summary_col = col['source'] == 'fiber_first' and key in ('union_prom', 'union_max')

    base_bg = _BLUE_LIGHT if is_summary_col else row_bg

    if thr_key:
        if col['source'] == 'event':
            val = ev.get(col['field'])
        else:
            val = fiber.get(col['field'])
        return _thr_bg(val, thr.get(thr_key)) or base_bg

    return base_bg


def _summary_value(col: dict, fiber: dict, delta_max: float | None) -> Any:
    key = col['key']
    fld = col['field']
    if key == 'delta_union_max':
        return delta_max
    if key == 'n_empalmes':
        return len(fiber.get('events', []))
    if fld == 'direction_label':
        return _DIR_LABELS.get(fiber.get('direction', 'normal'), fiber.get('direction'))
    return fiber.get(fld)


def _summary_bg(col: dict, fiber: dict, delta_max: float | None,
                thr: dict, row_bg: str) -> str:
    key = col['key']
    thr_key = col.get('thr_key')

    if key == 'delta_union_max' and delta_max is not None:
        return _RED_DELTA if delta_max > 0 else (_GREEN_DELTA if delta_max < 0 else row_bg)

    if thr_key:
        val = fiber.get(col['field'])
        return _thr_bg(val, thr.get(thr_key)) or row_bg

    return row_bg


# ── Sheet builders ────────────────────────────────────────────────────────

def _write_cable_sheet(wb: "Workbook", cable_name: str,
                       fibers: list[dict], thresholds: dict,
                       prev_snapshot: dict,
                       enabled_detail: list[dict]) -> None:
    clean = cable_name.strip()[:31]
    ws = wb.create_sheet(title=clean)
    ws.freeze_panes = "A7"
    ncols = len(enabled_detail)
    first = fibers[0] if fibers else {}

    # Header rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _cell(ws, 1, 1,
          value=f"Medición de Filamentos — Cable: {clean}",
          bold=True, bg=_BLUE_HDR, align="left", font_color="FFFFFF")
    ws.row_dimensions[1].height = 18

    info = (f"Fecha: {first.get('date','')}  |  "
            f"λ: {first.get('wavelength_nm','')} nm  |  "
            f"Equipo: {first.get('otdr_model','') or 'EXFO'}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    _cell(ws, 2, 1, value=info, align="left", bg=_BLUE_LIGHT)

    thr = thresholds
    thr_txt = (f"Umbrales:  Pérd. Unión > {thr.get('perdida_union_db',0.5)} dB  ·  "
               f"Atenuación > {thr.get('perdida_promedio_dbkm',0.25)} dB/km  ·  "
               f"Pérd. Intervalo > {thr.get('perdida_intervalo_db',2.0)} dB")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    _cell(ws, 3, 1, value=thr_txt, align="left", bg=_GRAY)
    ws.row_dimensions[4].height = 6

    # Column headers
    for ci, col in enumerate(enabled_detail, 1):
        _cell(ws, 6, ci, value=col['label'],
              bold=True, bg=_BLUE_SUB, font_color="FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = col['width']
    ws.row_dimensions[6].height = 30

    # Find index of 'fibra_num' column for merge
    fibra_col_idx = next(
        (ci for ci, c in enumerate(enabled_detail, 1) if c['key'] == 'fibra_num'),
        None)

    # Data rows
    row = 7
    for fiber in fibers:
        events = fiber.get('events', [])
        if not events:
            continue
        start_row = row
        row_bg = _GRAY if (fiber.get('fibra_num', 0) % 2 == 0) else _WHITE

        for ei, ev in enumerate(events):
            for ci, col in enumerate(enabled_detail, 1):
                val = _detail_value(col, fiber, ev, ei)
                bg  = _detail_bg(col, fiber, ev, ei, thr, row_bg)
                _cell(ws, row, ci, value=val, bg=bg, fmt=col.get('fmt'))
            row += 1

        if fibra_col_idx and len(events) > 1:
            ws.merge_cells(start_row=start_row, start_column=fibra_col_idx,
                           end_row=row - 1,   end_column=fibra_col_idx)


def _write_summary_sheet(wb: "Workbook", all_cables: dict[str, list[dict]],
                         thresholds: dict, previous: dict,
                         enabled_summary: list[dict]) -> None:
    ws = wb.create_sheet(title="Resumen")
    ws.freeze_panes = "A3"
    ncols = len(enabled_summary)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _cell(ws, 1, 1, value="Resumen General de Mediciones",
          bold=True, bg=_BLUE_HDR, font_color="FFFFFF", align="left")
    ws.row_dimensions[1].height = 18

    for ci, col in enumerate(enabled_summary, 1):
        _cell(ws, 2, ci, value=col['label'],
              bold=True, bg=_BLUE_SUB, font_color="FFFFFF")
        ws.column_dimensions[get_column_letter(ci)].width = col['width']
    ws.row_dimensions[2].height = 25

    row = 3
    thr = thresholds
    for cable_name, fibers in all_cables.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        _cell(ws, row, 1, value=f"  Cable: {cable_name.strip()}",
              bold=True, bg=_BLUE_LIGHT, align="left")
        row += 1

        prev_cable = previous.get(cable_name, {})
        for fiber in fibers:
            fibra_num = fiber.get('fibra_num', 0)
            direction = fiber.get('direction', 'normal')
            prev_key  = f"{fibra_num}_{direction}"
            prev_fiber = prev_cable.get(prev_key, {})
            delta_max = hist_mod.delta(
                fiber.get('perdida_union_maxima_db'),
                prev_fiber.get('perdida_union_maxima_db'))
            row_bg = _GRAY if fibra_num % 2 == 0 else _WHITE

            for ci, col in enumerate(enabled_summary, 1):
                val = _summary_value(col, fiber, delta_max)
                bg  = _summary_bg(col, fiber, delta_max, thr, row_bg)
                _cell(ws, row, ci, value=val, bg=bg, fmt=col.get('fmt'))
            row += 1


# ── Public API ────────────────────────────────────────────────────────────

def export_to_excel(cables_data: dict[str, list[dict]],
                    output_path: str,
                    thresholds: dict | None = None,
                    history_path: str = '',
                    column_config: dict | None = None) -> None:
    if not OPENPYXL_OK:
        raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

    thr = thresholds or {}
    col_cfg = column_config or default_column_config()

    enabled_detail  = active_cols(ALL_DETAIL_COLS,  col_cfg.get('detail',  []))
    enabled_summary = active_cols(ALL_SUMMARY_COLS, col_cfg.get('summary', []))

    history  = hist_mod.load(history_path) if history_path else {}
    ref_date = next(
        (f[0]['date'] for f in cables_data.values() if f and f[0].get('date')),
        None)
    previous = hist_mod.get_previous(history, ref_date)

    wb = Workbook()
    wb.remove(wb.active)

    for cable_name, fibers in cables_data.items():
        _write_cable_sheet(wb, cable_name, fibers, thr,
                           previous.get(cable_name, {}), enabled_detail)
    _write_summary_sheet(wb, cables_data, thr, previous, enabled_summary)

    wb.save(output_path)

    if history_path:
        hist_mod.save(history_path, cables_data, ref_date)


def build_output_path(root_folder: str, ref_date: str | None = None) -> str:
    ym = ref_date[:7] if ref_date and len(ref_date) >= 7 else str(date.today())[:7]
    return os.path.join(root_folder, f"FO_Cartilla_FOS_{ym}.xlsx")
